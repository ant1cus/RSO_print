import os
import re

from PyQt5.QtCore import QDir

import account_num

import openpyxl
from PyQt5.QtWidgets import QDialog, QMessageBox, QFileDialog
from openpyxl.utils import get_column_letter


class AccountNumWindow(QDialog, account_num.Ui_Dialog):  # Для файла с номерами
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.pushButton_cancel.clicked.connect(lambda: self.close())
        self.pushButton_create.clicked.connect(self.create_file)
        self.pushButton_path.clicked.connect(self.browse)

    def browse(self):  # Для кнопки открыть
        directory = QFileDialog.getExistingDirectory(self, "Find Files", QDir.currentPath())
        if directory:
            self.lineEdit_path.setText(directory)

    def create_file(self):
        path = self.lineEdit_path.text()
        if not path:
            QMessageBox.critical(self, 'УПС!', 'Путь к файлу учетных номеров пуст')
            return
        if not os.path.isdir(path):
            QMessageBox.critical(self, 'УПС!', 'Указанный путь к файлу учетных номеров не является директорией')
            return
        number_start = " ".join(self.lineEdit_account_num_start.text().split())
        if not number_start:
            QMessageBox.critical(self, 'УПС!', 'Нет начального учетного номера')
            return
        for el in number_start:
            if re.match(r'[A-Za-z0-9\s]', el):
                pass
            else:
                QMessageBox.critical(self, 'УПС!', 'Некорректные символы в учетном номере')
                return
        wb = openpyxl.Workbook()  # Открываем книгу
        ws = wb.active  # Активный лист
        # Начальные номера
        i = 0
        j = 1
        ws.column_dimensions[get_column_letter(j)].width = 12  # Ширина столбцов для отображения
        for el in range(1, 25001):  # Заполняем
            i += 1
            ws.cell(i, j).value = number_start + '/' + str(el)
            if i % 100 == 0:  # Переходим на следующий столбец, чтобы немного значений в одном
                i = 0
                j += 1
                ws.column_dimensions[get_column_letter(j)].width = 12  # Ширина
        wb.save(path + '/' + 'Файл учетных номеров № ' + number_start + '.xlsx')  # Сохраняем
        wb.close()  # Закрываем
        self.close()  # Закрываем окно
