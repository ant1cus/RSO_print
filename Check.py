import os
import pathlib
import re
import time

import openpyxl
import pandas as pd
import psutil


def doc_format(lineedit_old, lineedit_new, lineedit_file_num, radiobutton_fsb_df, radiobutton_fstek_df,
               combobox_classified, lineedit_num_scroll, lineedit_list_item, lineedit_number, checkbox_add_list_item,
               lineedit_add_list_item, lineedit_protocol,
               lineedit_conclusion, lineedit_prescription, lineedit_print, lineedit_executor_acc_sheet, label_protocol,
               label_conclusion, label_prescription, label_print, label_executor_acc_sheet, dateedit_date, lineedit_act,
               lineedit_statement, checkbox_conclusion_number, lineedit_conclusion_number,
               dateedit_conclusion_number,
               groupbox_inventory_insert, radiobutton_40_num, radiobutton_all_doc,
               lineedit_account_post, lineedit_account_signature, lineedit_account_path, hdd_number,
               groupbox_form27_insert, lineedit_firm, lineedit_path_form_27_create, qroupbox_instance,
               lineedit_number_instance, checkbox_conclusion, checkbox_protocol, checkbox_preciption, package,
               action_mo, groupbox_sp, lineedit_path_folder_sp, checkbox_name_gk, lineedit_name_gk,
               checkbox_conclusion_sp, checkbox_protocol_sp, checkbox_preciption_sp, checkbox_infocard_sp,
               lineedit_path_file_sp, checkbox_file_num):
    def check(n, e):
        for symbol in e:
            if n == symbol:
                return False
        return True

    for proc in psutil.process_iter():
        if proc.name() == 'WINWORD.EXE':
            return ['УПС!', 'Закройте все файлы Word!']
    answer = {'package': True if package.isChecked() else False, 'action_MO': True if action_mo.isChecked() else False,
              'path_old': lineedit_old.text().strip()}
    # answer['package'] = True if package.isChecked() else False
    # answer['action_MO'] = True if action_mo.isChecked() else False
    # Путь к исходным документам и проверки
    if not answer['path_old']:
        return ['УПС!', 'Путь к исходным документам пуст']
    if os.path.exists(answer['path_old']) is False:
        return ['УПС!', 'Папка с исходными документами отсутствует или переименована']
    if os.path.isdir(answer['path_old']):
        if len(os.listdir(answer['path_old'])) == 0:
            return ['УПС!', 'Папка с исходными документами пуста']
        if answer['package']:
            docs = []
            for folder in os.listdir(answer['path_old']):
                if os.path.isdir(answer['path_old'] + '\\' + folder):
                    # Ошибка если есть файлы старого формата
                    docs += [i for i in os.listdir(answer['path_old'] + '\\' + folder) if i[-3:] == 'doc']
        else:
            error = [i for i in os.listdir(answer['path_old']) if os.path.isdir(answer['path_old'] + '\\' + i) and
                     ('материалы' not in i.lower())]
            if error:
                return ['УПС!', 'В директории для преобразования присутствуют папки']
            # Ошибка если есть файлы старого формата
            docs = [i for i in os.listdir(answer['path_old']) if i[-3:] == 'doc']
        if len(docs) != 0:
            text = 'Файлы старого формата:\n' + '\n'.join(docs)
            return ['УПС!', text]
    else:
        return ['УПС!', 'Указанный путь к исходным документам не является директорией']
    if answer['action_MO']:
        if answer['package']:
            for folder_mo in os.listdir(answer['path_old']):
                file_mo = [mo for mo in os.listdir(answer['path_old'] + '\\' + folder_mo)
                           if mo[-3:] == 'txt' and 'F19' in mo]
                if not file_mo:
                    return ['УПС!', f'Нет текстового файла с серийниками для создания отчёта для'
                                    f' МО в папке {folder_mo}']
        else:
            file_mo = [mo for mo in os.listdir(answer['path_old']) if mo[-3:] == 'txt' and 'F19' in mo]
            if not file_mo:
                return ['УПС!', 'Нет текстового файла с серийниками для создания отчёта для МО']
    # Путь к конечным документам и проверки
    answer['path_new'] = lineedit_new.text().strip()
    if os.path.exists(answer['path_new']) is False:
        return ['УПС!', 'Папка с конечными документами отсутствует или переименована']
    if not answer['path_new']:
        return ['УПС!', 'Путь к конечной папке пуст']
    if os.path.isfile(answer['path_new']):
        return ['УПС!', 'Указанный путь к конечной папке не является директорией']
    if len(os.listdir(answer['path_new'])) != 0:
        return ['УПС!', 'Конечная папка не пуста, очистите директорию']
    answer['file_num'] = False
    if checkbox_file_num.isChecked():
        answer['file_num'] = lineedit_file_num.text().strip()
        if answer['file_num']:
            if os.path.isdir(answer['file_num']):
                return ['УПС!', 'Указанный путь к файлу номеров является директорией']
            else:
                if os.path.exists(answer['file_num']):
                    if answer['file_num'].endswith('.xlsx') is False:
                        return ['УПС!', 'Файл номеров не формата .xlsx']
                else:
                    return ['УПС!', 'Файл номеров удалён или переименован']
        else:
            return ['УПС!', 'Не указан файл номеров']
        file_in_directory = []
        if answer['package']:
            for directory in os.listdir(answer['path_old']):
                file_in_directory = [file for file in os.listdir(pathlib.Path(answer['path_old'], directory))
                                     if file.endswith('.docx')]
        else:
            file_in_directory = [file for file in os.listdir(answer['path_old']) if file.endswith('.docx')]
        dict_file = {}
        wb = openpyxl.load_workbook(answer['file_num'])  # Откроем книгу.
        ws = wb.active  # Делаем активным первый лист.
        for i in range(1, ws.max_row + 1):  # Пока есть значения
            if ws.cell(i, 1).value:
                dict_file[ws.cell(i, 1).value] = [ws.cell(i, 2).value + 'c',
                                                  ws.cell(i, 3).value.strftime("%d.%m.%Y")]  # Делаем список
        error_for_file_num = []
        for file in file_in_directory:
            # accepted_file = False
            # for name_file in ['акт', 'заключение', 'протокол', 'предписание', 'сопроводит', 'опись']:
            #     if name_file in file.lower():
            accepted_file = [False if name_file in file.lower() else True for name_file in ['акт', 'заключение', 'протокол', 'предписание', 'сопроводит', 'опись']]
            if all(accepted_file):
                continue
            num_date = dict_file.pop(file.rpartition('.')[0], 'File not found')
            if num_date == 'File not found':
                error_for_file_num.append(f'Документ {file} не найден в файле номеров')
            else:
                if num_date[0] is False:
                    error_for_file_num.append(f'Для записи {file} в файле номеров не указан секретный номер')
                elif num_date[1] is False:
                    error_for_file_num.append(f'Для записи {file} в файле номеров не указана дата')
        if dict_file:
            for file in dict_file:
                # error_for_file_num.append(f'В файле номеров указан документ {file}, которого нет в исходных файлах')
                if dict_file[file][0] is False:
                    error_for_file_num.append(f'Для записи {file} в файле номеров не указан секретный номер')
                elif dict_file[file] is False:
                    error_for_file_num.append(f'Для записи {file} в файле номеров не указана дата')
        if error_for_file_num:
            return ['УПС!', '\n'.join(error_for_file_num)]
    answer['path_sp'], answer['path_file_sp'], answer['name_gk'], answer['check_sp'] = False, False, False, []
    if groupbox_sp.isChecked():
        answer['path_sp'] = lineedit_path_folder_sp.text().strip()
        if not answer['path_sp']:
            return ['УПС!', 'Путь к папке с материалами СП пуст']
        if os.path.isfile(answer['path_sp']):
            return ['УПС!', 'Указанный путь к материалам СП не является директорией']
        answer['path_file_sp'] = lineedit_path_file_sp.text().strip()
        if not answer['path_file_sp']:
            return ['УПС!', 'Путь к файлу с номерами СП пуст']
        if os.path.isdir(answer['path_file_sp']):
            return ['УПС!', 'Указанный путь к файлу с номерами СП не является файлом']
        if os.path.exists(answer['path_file_sp']) is False:
            return ['УПС!', 'Файл номеров СП удалён или переименован']
        if answer['path_file_sp'].endswith('.xlsx') is False:
            return ['УПС!', 'Файл номеров СП не формата .xlsx']
        if checkbox_name_gk.isChecked():
            answer['name_gk'] = lineedit_name_gk.text().strip()
            if answer['name_gk'] is False:
                return ['УПС!', 'Введите имя ГК']
        if len(os.listdir(answer['path_sp'])) == 0 and answer['name_gk'] is False:
            return ['УПС!', 'Папка с материалами СП пуста (введите имя ГК или добавьте материалы)']
        answer['check_sp'] = [True if i.isChecked() else False for i in [checkbox_conclusion_sp, checkbox_protocol_sp,
                                                                         checkbox_preciption_sp, checkbox_infocard_sp]]
        if all(i is False for i in answer['check_sp']):
            return ['УПС!', 'Не выбран ни один документ для проверки СП']
    # Ведомство
    if radiobutton_fsb_df.isChecked():
        answer['service'] = True
    elif radiobutton_fstek_df.isChecked():
        answer['service'] = False
    else:
        return ['УПС!', 'Не выбрано ведомство для вставки колонтитулов']
    # Гриф секретности
    class_ = {'ДСП': 'Для служебного пользования', 'С': 'Секретно', 'СС': 'Совершенно секретно',
              'ОВ': 'Особой важности'}
    classified = combobox_classified.currentText().strip()
    if not classified:
        return ['УПС!', 'Не выбрана категория секретности']
    answer['classified'] = class_[classified]
    # Номер экземпляра
    answer['num_scroll'] = lineedit_num_scroll.text().strip()
    if not answer['num_scroll']:
        return ['УПС!', 'Не указан номер экземпляра']
    # Пункт перечня
    answer['list_item'] = lineedit_list_item.text().strip()
    if not answer['list_item']:
        return ['УПС!', 'Не указан пункт перечня']
        # Дополнительный пункт перечня
    answer['add_list_item'] = False
    if checkbox_add_list_item.isChecked():
        answer['add_list_item'] = lineedit_add_list_item.text().strip()
        if not answer['add_list_item']:
            return ['УПС!', 'Не указан дополнительный пункт перечня']
    # Номер
    answer['number'] = lineedit_number.text().strip()
    if answer['number']:
        if answer['number'][-1] in ['С', 'с']:
            answer['number'] = answer['number'].replace(answer['number'][-1], 'c')
        if not answer['number']:
            return ['УПС!', 'Не указан номер']
        for i in answer['number']:
            if check(i, ('1', '2', '3', '4', '5', '6', '7', '8', '9', '0', '/', 'c', 'с', '-', 'Н', 'С', 'с')):
                return ['УПС!', 'Есть лишние символы в номере']
        if (re.match(r'\w+/\w+/\w+c$', answer['number']) is None) and (re.match(r'НС-\w+c$', answer['number']) is None):
            return ['УПС!', 'Секретный номер указан неверно']
    else:
        return ['УПС!', 'Не указан секретный номер']
    # Исполнитель, заключение, предписание, протокол, печать
    answer['act'] = lineedit_act.text().strip()
    answer['statement'] = lineedit_statement.text().strip()
    answer['protocol'] = lineedit_protocol.text().strip()
    answer['conclusion'] = lineedit_conclusion.text().strip()
    answer['prescription'] = lineedit_prescription.text().strip()
    answer['print_people'] = lineedit_print.text().strip()
    answer['executor_acc_sheet'] = lineedit_executor_acc_sheet.text().strip()
    list_label = [label_protocol, label_conclusion, label_prescription, label_print, label_executor_acc_sheet]
    i = 0
    for element in [answer['protocol'], answer['conclusion'], answer['prescription'],
                    answer['print_people'], answer['executor_acc_sheet']]:
        if not element:
            return ['УПС!', 'Не указан(а) ' + list_label[i].text()]
        i += 1
    # Дата
    answer['date'] = dateedit_date.date().toString('dd.MM.yyyy')
    # answer['date'] = lineedit_date.text().strip()
    # try:
    #     time.strptime(answer['date'], '%d.%m.%Y')
    # except ValueError:
    #     return ['УПС!', 'Формат даты указан неверно! (необходимый формат: dd.mm.yyyy)']
    answer['conclusion_number'] = False
    answer['conclusion_number_date'] = False
    if checkbox_conclusion_number.isChecked():
        answer['conclusion_number'] = lineedit_conclusion_number.text().strip()
        if answer['conclusion_number'][-1] in ['С', 'с']:
            answer['conclusion_number'] = answer['conclusion_number'].replace(answer['conclusion_number'][-1], 'c')
        if not answer['conclusion_number']:
            return ['УПС!', 'Не указан номер заключения']
        for i in answer['conclusion_number']:
            if check(i, ('1', '2', '3', '4', '5', '6', '7', '8', '9', '0', '/', 'c', 'с', '-', 'Н', 'С', 'с')):
                return ['УПС!', 'Есть лишние символы в номере заключения']
        if (re.match(r'\w+/\w+/\w+c$', answer['conclusion_number']) is None)\
                and (re.match(r'НС-\w+c$', answer['conclusion_number']) is None):
            return ['УПС!', 'Номер заключения указан неверно']
        answer['conclusion_number_date'] = dateedit_conclusion_number.date().toString('dd.MM.yyyy')
        # answer['conclusion_number_date'] = lineedit_add_conclusion_number_date.text().strip()
        # try:
        #     time.strptime(answer['conclusion_number_date'], '%d.%m.%Y')
        # except ValueError:
        #     return ['УПС!', 'Формат доп. даты заключения указан неверно! (необходимый формат: dd.mm.yyyy)']
    answer['account'], answer['flag_inventory'], answer['account_post'] = False, False, False
    answer['account_signature'], answer['account_path'] = False, False
    if groupbox_inventory_insert.isChecked():
        answer['account'] = True
        if radiobutton_40_num.isChecked() or radiobutton_all_doc.isChecked():
            answer['flag_inventory'] = 40 if radiobutton_40_num.isChecked() else 1
        else:
            return ['УПС!', 'Не указано количество документов в описе']
        answer['account_post'] = lineedit_account_post.text().strip()
        if not answer['account_post']:
            return ['УПС!', 'Не указана должность для описи']
        answer['account_signature'] = lineedit_account_signature.text().strip()
        if not answer['account_signature']:
            return ['УПС!', 'Не указана подпись для описи']
        answer['account_path'] = lineedit_account_path.text().strip()
        if not answer['account_path']:
            return ['УПС!', 'Не указан путь для файла описи']
        else:
            if os.path.isdir(answer['account_path']):
                pass
            else:
                return ['УПС!', 'Для описи необходимо указать файл']
    if not hdd_number:
        return ['УПС!', 'Отсутствует номер жесткого диска']
    else:
        answer['hdd_number'] = hdd_number
    answer['firm'] = False
    answer['form_27'] = False
    if groupbox_form27_insert.isChecked():
        answer['firm'] = lineedit_firm.text().strip()
        if not answer['firm']:
            return ['УПС!', 'Не заполнена организация для 27 формы']
        answer['form_27'] = lineedit_path_form_27_create.text().strip()
        if not answer['form_27']:
            return ['УПС!', 'Нет пути для 27 формы']
        else:
            if os.path.isdir(answer['form_27']):
                pass
            else:
                return ['УПС!', 'Указанный путь для 27 формы не является директорией']
    answer['second_copy'] = []
    answer['number_instance'] = []
    if qroupbox_instance.isChecked():
        number_instance = lineedit_number_instance.text().strip()
        if not number_instance:
            return ['УПС!', 'Не указаны номера экземпляров']
        for i in number_instance:
            if check(i, ('1', '2', '3', '4', '5', '6', '7', '8', '9', '0', ' ', '-', ',', '.')):
                return ['УПС!', 'Есть лишние символы в номерах экземпляров']
        set_num = number_instance.replace(' ', '').replace(',', '.')
        if set_num[0] == '.' or set_num[0] == '-':
            return ['УПС!', 'Первый символ введён не верно']
        if set_num[-1] == '.' or set_num[-1] == '-':
            return ['УПС!', 'Последний символ введён не верно']
        for i in range(len(set_num)):
            if set_num[i] == '.' or set_num[i] == '-':
                if set_num[i + 1] == '.' or set_num[i + 1] == '-':
                    return ['УПС!', 'Два разделителя номеров подряд']
        set_number = []
        for element in set_num.split('.'):
            if '-' in element:
                num1, num2 = int(element.partition('-')[0]), int(element.partition('-')[2])
                if num1 >= num2:
                    return ['УПС!', 'Диапазон номеров экземпляров указан не верно']
                else:
                    for el in range(num1, num2 + 1):
                        set_number.append(el)
            else:
                set_number.append(element)
        set_number.sort()
        if len(set_number) != len(set(set_number)):
            return ['УПС!', 'Есть повторения в номерах экземпляров']
        answer['number_instance'] = set_number
        answer['second_copy'] = [True if i.isChecked() else False for i in [checkbox_conclusion, checkbox_protocol,
                                                                            checkbox_preciption]]
        if all(i is False for i in answer['second_copy']):
            return ['УПС!', 'Не выбран ни один документ для создания экземпляров']
    return answer


def doc_print(radiobutton_fsb_print, radiobutton_fstek_print, checkbox_conclusion_print, checkbox_protocol_print,
              checkbox_preciption_print, lineedit_old_print, lineedit_account_numbers,
              checkbox_add_account_numbers, lineedit_add_account_numbers, checkbox_form_27, lineedit_path_form_27_print,
              button_gr, lineedit_printer, checkbox_print_order, path_for_default, package):
    answer = {'path_for_default': path_for_default, 'package_': True if package.isChecked() else False,
              'print_order': True if checkbox_print_order.isChecked() else False}
    # Ведомство
    if radiobutton_fsb_print.isChecked():
        answer['service'] = True
    elif radiobutton_fstek_print.isChecked():
        answer['service'] = False
    else:
        return ['УПС!', 'Не выбрано ведомство при печати документов']
    answer['document_list'] = {i: True if j.isChecked() else False for i, j in zip(['заключение', 'протокол',
                                                                                    'предписание'],
                                                                                   [checkbox_conclusion_print,
                                                                                    checkbox_protocol_print,
                                                                                    checkbox_preciption_print])}
    answer['path_old_print'] = lineedit_old_print.text().strip()
    if not answer['path_old_print']:
        return ['УПС!', 'Путь к исходным документам для печати отсутствует или переименован']
    if not answer['path_old_print']:
        return ['УПС!', 'Путь к исходным документам для печати пуст']
    if os.path.isdir(answer['path_old_print']):
        if len(os.listdir(answer['path_old_print'])) == 0:
            return ['УПС!', 'Папка с исходными документами для печати пуста']
        docs = []
        if answer['package_']:
            error = [i for i in os.listdir(answer['path_old_print'])
                     if os.path.isfile(pathlib.Path(answer['path_old_print'], i))]
            print(error)
            if error:
                return ['УПС!', 'В директории для пакетной печати присутствуют файлы']
            for folder in os.listdir(answer['path_old_print']):
                # Ошибка если есть файлы старого формата
                docs = docs + [i for i in os.listdir(pathlib.Path(answer['path_old_print'], folder)) if i[-3:] == 'doc']
        else:
            error = [i for i in os.listdir(answer['path_old_print'])
                     if os.path.isdir(pathlib.Path(answer['path_old_print'], i))]
            if error:
                return ['УПС!', 'В директории для преобразования присутствуют папки']
            # Ошибка если есть файлы старого формата
            docs = [i for i in os.listdir(answer['path_old_print']) if i[-3:] == 'doc']
        if len(docs) != 0:
            text = 'Файлы старого формата:\n' + '\n'.join(docs)
            return ['УПС!', text]
    else:
        return ['УПС!', 'Указанный путь к исходным документам для печати не является директорией']
    # Путь к номерам
    answer['path_account_num'] = lineedit_account_numbers.text().strip()
    answer['add_path_account_num'] = False
    if checkbox_add_account_numbers.isChecked():
        answer['add_path_account_num'] = lineedit_add_account_numbers.text().strip()
        if not answer['add_path_account_num']:
            return ['УПС!', 'Путь к доп. файлу номеров учетных листов пуст']
        if os.path.isdir(answer['add_path_account_num']):
            return ['УПС!', 'Указанный путь к доп. файлу номеров учётных листов является директорией']
        else:
            if os.path.exists(answer['add_path_account_num']):
                if answer['add_path_account_num'].endswith('.xlsx'):
                    pass
                else:
                    return ['УПС!', 'Доп. файл номеров не формата .xlsx']
            else:
                return ['УПС!', 'Доп. файл номеров удалён или переименован']
    if not answer['path_account_num']:
        return ['УПС!', 'Путь к файлу номеров учетных листов пуст']
    if os.path.isdir(answer['path_account_num']):
        return ['УПС!', 'Указанный путь к файлу номеров учётных листов является директорией']
    else:
        if os.path.exists(answer['path_account_num']):
            if answer['path_account_num'].endswith('.xlsx'):
                try:
                    df_acc_num = pd.read_excel(answer['path_account_num'], header=None)
                    if df_acc_num.empty:
                        return ['УПС!', 'Файл номеров пустой']
                except BaseException:
                    return ['УПС!', 'Что-то не так с файлом номеров']
            else:
                return ['УПС!', 'Файл номеров не формата .xlsx']
        else:
            return ['УПС!', 'Файл номеров удалён или переименован']
    # Форма 27
    answer['path_form_27'] = False
    if checkbox_form_27.isChecked():
        if answer['package_'] is False:
            answer['path_form_27'] = lineedit_path_form_27_print.text().strip()
            if not answer['path_form_27']:
                return ['УПС!', 'Путь к 27 форме пуст']
            if os.path.isdir(answer['path_form_27']):
                return ['УПС!', 'Указанный путь к 27 форме является директорией']
            else:
                if os.path.exists(answer['path_form_27']) and not answer['package_']:
                    if answer['path_form_27'].endswith('.xlsx'):
                        pass
                    else:
                        return ['УПС!', 'Файл "Форма 27" не формата .xlsx']
                else:
                    if not answer['package_']:
                        return ['УПС!', 'Файл "Форма 27" удалён или переименован']
        else:
            answer['path_form_27'] = True
    # Способ печати
    answer['print_flag'] = False
    for button in button_gr:
        if button.isChecked():
            answer['print_flag'] = button.text()
    if not answer['print_flag']:
        return ['УПС!', 'Не указан метод печати']
    answer['name_printer'] = lineedit_printer.text().strip()
    if not answer['name_printer']:
        return ['УПС!', 'Не выбран принтер']
    return answer
