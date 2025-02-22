import datetime
import os
import re
import shutil
import time
import traceback
import zipfile
import itertools
import threading
import docx
import fitz
import numpy
import numpy as np
import openpyxl
import pythoncom
import pandas as pd
import openpyxl.styles
from small_functions import sorting_files, return_error, pages_count
from get_text_num import create_text_for_docs
from pathlib import Path
from DoingWindow import DoWindow
from itertools import groupby
from PyQt5.QtCore import QThread, pyqtSignal
from docx.enum.section import WD_ORIENTATION
from docx.enum.text import WD_PARAGRAPH_ALIGNMENT
from docx.oxml import OxmlElement, ns
from docx.shared import Pt
from openpyxl.utils import get_column_letter
from natsort import natsorted
from word2pdf import word2pdf
from lxml import etree
from format_documets import format_doc


class CancelException(Exception):
    pass


class FormatDoc(QThread):  # Если требуется вставить колонтитулы
    status_finish = pyqtSignal(str, str)
    progress_value = pyqtSignal(int)
    info_value = pyqtSignal(str, str)
    status = pyqtSignal(str)
    line_progress = pyqtSignal(str)
    line_doing = pyqtSignal(str)

    def __init__(self, incoming_data):  # Список переданных элементов.
        QThread.__init__(self)
        self.path_old = incoming_data['path_old']
        self.path_new = incoming_data['path_new']
        self.file_num = incoming_data['file_num']
        self.classified = incoming_data['classified']
        self.num_scroll = incoming_data['num_scroll']
        self.list_item = incoming_data['list_item']
        self.number = incoming_data['number']
        self.protocol = incoming_data['protocol']
        self.conclusion = incoming_data['conclusion']
        self.prescription = incoming_data['prescription']
        self.print_people = incoming_data['print_people']
        self.date = incoming_data['date']
        self.executor_acc_sheet = incoming_data['executor_acc_sheet']
        self.account = incoming_data['account']
        self.flag_inventory = incoming_data['flag_inventory']
        self.account_post = incoming_data['account_post']
        self.account_signature = incoming_data['account_signature']
        self.account_path = incoming_data['account_path']
        self.firm = incoming_data['firm']
        self.path_form_27 = incoming_data['form_27']
        self.second_copy = incoming_data['second_copy']
        self.service = incoming_data['service']
        self.hdd_number = incoming_data['hdd_number']
        self.q = incoming_data['queue']
        self.logging = incoming_data['logging']
        self.package = incoming_data['package']
        self.report_rso = incoming_data['action_MO']
        self.act = incoming_data['act']
        self.statement = incoming_data['statement']
        self.number_instance = incoming_data['number_instance']
        self.path_sp = incoming_data['path_sp']
        self.path_file_sp = incoming_data['path_file_sp']
        self.name_gk = incoming_data['name_gk']
        self.check_sp = incoming_data['check_sp']
        self.conclusion_number = incoming_data['conclusion_number']
        self.conclusion_number_date = incoming_data['conclusion_number_date']
        self.add_list_item = incoming_data['add_list_item']
        self.num_1 = self.num_2 = 0
        self.event = threading.Event()
        self.event.set()
        self.move = incoming_data['move']
        self.default_path = incoming_data['default_path']
        self.all_doc = 0
        self.now_doc = 0
        self.percent_progress = 0
        self.name_dir = Path(self.path_old).name
        title = f'Регистрация документов в папке «{self.name_dir}»'
        self.doing_window = DoWindow(self.default_path, self.event, self.move, title)
        self.progress_value.connect(self.doing_window.progressBar.setValue)
        self.line_progress.connect(self.doing_window.lineEdit_progress.setText)
        self.line_doing.connect(self.doing_window.lineEdit_doing.setText)
        self.info_value.connect(self.doing_window.info_message)
        self.doing_window.show()

    def run(self):
        time_start = datetime.datetime.now()
        # self.progress.emit(0)  # Обновляем статус бар
        self.status.emit('Начинаем')
        self.logging.info("Старт программы")
        pt_num = 14 if self.service else 12
        dict_40 = []  # Список документов в опись
        for_27 = []
        try:  # Для отлова ошибок
            if self.file_num:  # Если есть файл номеров
                dict_file = {}
                wb = openpyxl.load_workbook(self.file_num)  # Откроем книгу.
                ws = wb.active  # Делаем активным первый лист.
                for i in range(1, ws.max_row + 1):  # Пока есть значения
                    if ws.cell(i, 1).value:
                        dict_file[ws.cell(i, 1).value] = [ws.cell(i, 2).value + 'c',
                                                          ws.cell(i, 3).value.strftime("%d.%m.%Y")]  # Делаем список
            else:  # Если нет файла номеров
                if re.match(r'\w+/\w+/\w+c', self.number):
                    self.num_1 = self.number.rpartition('/')[0] + '/'
                    self.num_2 = self.number.rpartition('/')[2].rpartition('c')[0]
                else:
                    self.num_1 = self.number.partition('-')[0] + '-'
                    self.num_2 = self.number.partition('-')[2].rpartition('c')[0]
            self.logging.info("Созданы секретные номера")
            errors = []
            # Подсчёт кол-ва документов
            # Сначала номера, потом в 27 форму
            columns_name = ['name', 'number', 'start_path', 'finish_path', 'parent_path', 'secret_number', 'action',
                            'conclusion', 'protocol', 'prescription', 'text', 'change_date', 'executor',
                            'first_header_text', 'footer_text', 'date', 'text_conclusion', 'text_prescription',
                            'text_finish', 'pages']
            documents = pd.DataFrame(columns=columns_name)
            for item in Path(self.path_old).rglob('*.*'):
                documents = documents.append({'name': item.name, 'start_path': item, 'parent_path': item.parent,
                                              'print_people': self.print_people, 'date': self.date}, ignore_index=True)
            # documents = {item: {'name': item.name, 'start_path': item} for item in Path(self.path_old).rglob('*.*')}
            folders = [item for item in Path(self.path_old).glob('*') if os.path.isdir(item)]
            docs_txt = [file for file in os.listdir(self.path_old) if file.endswith('.txt')]  # Список txt
            if self.package:
                for folder in folders:
                    finish_path = Path(self.path_new, folder.name)
                    Path(finish_path).mkdir(parents=True, exist_ok=True)
                    # while True:
                    #     if os.listdir(finish_path):
                    #         self.info_value.emit('Вопрос?', f'В конечной папке «{folder.name}» уже присутствует файлы.'
                    #                                         f' Удалите или переместите её.')
                    #         self.event.clear()
                    #         self.event.wait()
                    #         if self.doing_window.stop_threading:
                    #             return error_return(self.logging, 'Прервано пользователем', self.status,
                    #                                 'Прервано пользователем', self.default_path,
                    #                                 [self.status_finish, 'format_doc', str(self)], self.doing_window)
                    #     else:
                    #         Path(finish_path).mkdir(parents=True, exist_ok=True)
                    #         break
                    # Надо посмтореть что там с ФСО
                    for file in folder.rglob('*.*'):
                        documents.loc[documents.loc[documents['start_path'] == file].index[0], 'finish_path'] \
                            = Path(self.path_new, file.name)
                        # documents[file]['finish_path'] = Path(finish_path, file.name)
                    files_order = sorting_files(folder, self.path_new, False)
                    if files_order['error']:
                        return return_error(self.logging,
                                            f"Регистрация документов в папке «{self.name_dir}» "
                                            f"не завершена из-за ошибки",
                                            self.status, f"Ошибка при регистрации документов в папке «{self.name_dir}»",
                                            self.default_path, [self.status_finish, 'format_doc', str(self)],
                                            self.doing_window,
                                            self.event, files_order['text'],
                                            [self.info_value, 'УПС!', files_order['text']])
                    self.percent_progress += files_order['data']['docs_for_progress']
                    answer = create_text_for_docs(self.logging, files_order['data']['docs'], documents, self.num_1,
                                                  self.num_2, self.conclusion_number, self.conclusion_number_date,
                                                  self.date, self.conclusion, self.protocol, self.prescription,
                                                  self.print_people, self.act,
                                                  self.statement, self.classified, self.num_scroll, self.list_item,
                                                  self.hdd_number, dict_40, for_27, self.line_doing, self.add_list_item,
                                                  self.file_num)
                    if answer['error']:
                        return return_error(self.logging,
                                            f"Регистрация документов в папке «{self.name_dir}» не завершена из-за ошибки",
                                            self.status, f"Ошибка при регистрации документов в папке «{self.name_dir}»",
                                            self.default_path, [self.status_finish, 'format_doc', str(self)],
                                            self.doing_window,
                                            self.event, answer['text'], [self.info_value, 'УПС!', answer['text']])
                    self.num_1, self.num_2 = answer['data']['num_1'], answer['data']['num_2']
            else:
                for file in Path(self.path_old).rglob('*.*'):
                    documents.loc[documents.loc[documents['start_path'] == file].index[0], 'finish_path']\
                        = Path(self.path_new, file.name)
                    # documents[file]['finish_path'] = Path(self.path_new, file.name)
                files_order = sorting_files(self.path_old, self.path_new, False)
                if files_order['error']:
                    return return_error(self.logging,
                                        f"Регистрация документов в папке «{self.name_dir}» не завершена из-за ошибки",
                                        self.status, f"Ошибка при регистрации документов в папке «{self.name_dir}»",
                                        self.default_path, [self.status_finish, 'format_doc', str(self)],
                                        self.doing_window,
                                        self.event, files_order['text'], [self.info_value, 'УПС!', files_order['text']])
                self.percent_progress += files_order['data']['docs_for_progress']
                answer = create_text_for_docs(self.logging, files_order['data']['docs'], documents, self.num_1,
                                              self.num_2, self.conclusion_number, self.conclusion_number_date,
                                              self.date, self.conclusion, self.protocol, self.prescription,
                                              self.print_people, self.act,
                                              self.statement, self.classified, self.num_scroll, self.list_item,
                                              self.hdd_number, dict_40, for_27, self.line_doing, self.add_list_item,
                                              self.file_num)

                if answer['error']:
                    return return_error(self.logging,
                                        f"Регистрация документов в папке «{self.name_dir}» не завершена из-за ошибки",
                                        self.status, f"Ошибка при регистрации документов в папке «{self.name_dir}»",
                                        self.default_path, [self.status_finish, 'format_doc', str(self)],
                                        self.doing_window,
                                        self.event, answer['text'], [self.info_value, 'УПС!', answer['text']])
                self.num_1, self.num_2 = answer['data']['num_1'], answer['data']['num_2']

                for document in documents.iterrows():
                    print(document)
            print()
            # if self.package:
            #     for folder in os.listdir(self.path_old):
            #         self.progress.emit(0)  # Начинаем прогресс бар для каждой папки
            #         if os.path.isdir(self.path_old + '\\' + folder):
            #             os.chdir(self.path_old + '\\' + folder)
            #             path_old = self.path_old + '\\' + folder
            #             path = self.path_new + '\\' + folder
            #             try:
            #                 os.mkdir(path)
            #             except FileExistsError:
            #                 self.status.emit('Ошибка')  # Сообщение в статус бар
            #                 self.messageChanged.emit('УПС!', 'В конечной папке уже присутствует папка «'
            #                                          + path.rpartition('\\')[2] + '». Удалите или переместите её.')
            #                 return
            #         elif os.path.isfile(self.path_old + '\\' + folder):
            #             self.logging.info(folder + ' является файлом, пропускаем')
            #             continue
            #         else:
            #             os.chdir(self.path_old)
            #             path_old = self.path_old
            #             path = self.path_new
            #         return_val = format_doc(path_old, self.classified, self.list_item, self.num_scroll, self.account,
            #                                  self.firm, self.logging, self.status, path, self.file_num, self.num_1,
            #                                  self.num_2, self.date, self.conclusion, self.protocol, self.prescription,
            #                                  self.hdd_number, self.print_people, self.progress, self.flag_inventory,
            #                                  self.account_post, self.account_signature, path,
            #                                  self.executor_acc_sheet, self.service, False, self.number_instance,
            #                                  self.path_sp, self.name_gk, self.check_sp, self.conclusion_number,
            #                                  self.conclusion_number_date)
            #         if return_val['error']:
            #             self.logging.info("Конец программы, время работы: " + str(datetime.datetime.now() - time_start))
            #             self.logging.info(
            #                 "\n*******************************************************************************\n")
            #             self.status.emit('Ошибки!')  # Посылаем значение если готово
            #             # self.progress.emit(0)  # Завершаем прогресс бар
            #             return
            #         # self.progress.emit(100)  # Завершаем прогресс бар
            #         self.num_1, self.num_2 = return_val['text'][0], return_val['text'][1]
            #         docs_txt = [file for file in os.listdir(path_old) if file[-4:] == '.txt']  # Список txt
            #         for txt_file in docs_txt:
            #             shutil.copy(txt_file, path)
            # else:
            #     return_val = format_doc(self.path_old, self.classified, self.list_item, self.num_scroll, self.account,
            #                              self.firm, self.logging, self.status, self.path_new, self.file_num, self.num_1,
            #                              self.num_2, self.date, self.conclusion, self.protocol, self.prescription,
            #                              self.hdd_number, self.print_people, self.progress, self.flag_inventory,
            #                              self.account_post, self.account_signature, self.account_path,
            #                              self.executor_acc_sheet, self.service, self.path_form_27, self.number_instance,
            #                              self.path_sp, self.name_gk, self.check_sp, self.conclusion_number,
            #                              self.conclusion_number_date)
            #     if return_val['error']:
            #         self.logging.info("Конец программы, время работы: " + str(datetime.datetime.now() - time_start))
            #         self.logging.info(
            #             "\n*******************************************************************************\n")
            #         self.status.emit('Ошибки!')  # Посылаем значение если готово
            #         # self.progress.emit(0)  # Завершаем прогресс бар
            #         return
            #     docs_txt = [file for file in os.listdir(self.path_old) if file[-4:] == '.txt']  # Список txt
            #     for txt_file in docs_txt:
            #         shutil.copy(txt_file, self.path_new)
            #     # self.progress.emit(100)  # Завершаем прогресс бар
            if errors:
                self.logging.warning('\n'.join(errors))
                self.logging.info("Конец программы, время работы: " + str(datetime.datetime.now() - time_start))
                self.logging.info("\n*******************************************************************************\n")
                self.status.emit('Завершено с ошибками!')  # Посылаем значение если готово
                # self.messageChanged.emit('ВНИМАНИЕ!', '\n'.join(errors))
            else:
                self.logging.info("Конец программы, время работы: " + str(datetime.datetime.now() - time_start))
                self.logging.info("\n*******************************************************************************\n")
                self.status.emit('Готово!')  # Посылаем значение если готово
        # except CancelException:
        #     self.logging.warning(f"Решистрация документов в папке «{self.name_dir}» отменена пользователем")
        #     self.status.emit(f"Регистрация документов в папке «{self.name_dir}» отменена пользователем")
        #     os.chdir(self.default_path)
        #     self.status_finish.emit('format_doc', str(self))
        #     time.sleep(1)  # Не удалять, не успевает отработать emit status_finish. Может потом
        #     self.window_check.close()
        #     return
        # except KeyError as keyError:  # Если ошибка по ключу
        #     self.status.emit('Ошибка')  # Сообщение в статус бар
        #     self.logging.error("Ошибка:\n " + str(keyError) + '\n' + traceback.format_exc())
        #     self.messageChanged.emit('УПС!', 'Программа не может найти файл ' + str(keyError))
            # self.progress.emit(0)  # Завершаем прогресс бар
        except BaseException as e:
            self.status.emit('Ошибка')  # Сообщение в статус бар
            self.logging.error("Ошибка:\n " + str(e) + '\n' + traceback.format_exc())
            # self.progress.emit(0)  # Завершаем прогресс бар
